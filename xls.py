
import openpyxl
from numpy import arange

def writecell(workbook, worksheet, row, column, content):
    """
    workbook  = xlsx filename
    worksheet = name
    row = number of the row as in the xlsx worksheet 
    column = str or index
    """
    #wb = workbook #openpyxl.load_workbook(workbook)
    #ws = wb.get_sheet_by_name(worksheet)
    wb = workbook
    ws = worksheet
    
    if type(column) == int:
        column = openpyxl.cell.get_column_letter(column)
    
    ws[column + str(row)] = content
    #wb.save(workbook)
    
def get_column_string_by_name(workbook, worksheet, column_name, row = 1):
    
    #wb = workbook #openpyxl.load_workbook(workbook)
    #ws = wb.get_sheet_by_name(worksheet)
    wb = workbook
    ws = worksheet
    
    max_col = ws.get_highest_column() 
    max_col = openpyxl.cell.get_column_letter(max_col)
    
    search_range = 'A'+str(row)+':'+max_col+str(row)
    
    for row in ws.iter_rows(search_range):
        for mycell in row:
            if mycell.value == column_name:
                xy = openpyxl.cell.coordinate_from_string(mycell.coordinate)
                return xy[0] 
            
def get_column_index_by_name(workbook, worksheet, column_name, row = 1):
    name = get_column_string_by_name(workbook, worksheet, column_name, row)
    return openpyxl.cell.column_index_from_string(name)
            
def get_row_index_by_name(workbook, worksheet, row_name, column = 'A'):
    
    #wb = workbook #openpyxl.load_workbook(workbook)
    #ws = wb.get_sheet_by_name(worksheet)
    wb = workbook
    ws = worksheet
    
    max_row = ws.max_row#get_highest_row() 
    #max_col = openpyxl.cell.get_column_letter(max_col)
    
    search_range = column+str(1)+':'+column+str(max_row)
    
    for row in ws.iter_rows(search_range):
        for mycell in row:
            if mycell.value == row_name:
                xy = openpyxl.utils.coordinate_from_string(mycell.coordinate)
                return xy[1] 

def load2dict(worksheet, dev_id):
    """
    :param worksheet: loaded worksheet of an xlrd workbook with one column 'ID' and other columns with more properties
        :code:
            workbook = xlrd.open_workbook('xls/A3172_S3_Measurements.xlsx')
            worksheet = workbook.sheet_by_index(0)
    :param dev_id: the name of the device for which will be searched in the column 'ID'
    :return: a dictionary with the names of the properties and the value of the device with the ID provided in 'dev_id'
    """
    column = worksheet.col_values(0) # ID

    ind0 = [i for i in arange(len(column)) if column[i] == 'ID']
    ind0 = int(ind0[0])
    keys = worksheet.row_values(ind0)

    ind = [i for i in arange(len(column)) if column[i] == dev_id]
    ind = int(ind[0])
    values = worksheet.row_values(ind)

    the_dict = {keys[i] : values[i] for i in arange(len(keys))}

    return the_dict
    
def get_value_in_column_by_other_columns(column_letter, other_row_values, other_column_letters, worksheet, return_row = False): 
    """ 
    input: 
        list of values expected in other columns 
        if this list of values is found the value in 'column_letter' is returned.
    
    usage example: 
        wb = openpyxl.load_workbook('test.xlsx', data_only = True)
        ws = wb.get_sheet_by_name('Sheet1')
        get_value_in_column_by_other_columns('C', ['A 6', 2.95], ['A', 'B'], ws)
    """
    other_column_letter_inds = [openpyxl.cell.column_index_from_string(x) for x in other_column_letters]
    column_letter_ind       = openpyxl.cell.column_index_from_string(column_letter)
    
    cell_value = None
    
    curr_row = 1
    
    while curr_row < 1000:

        cell_values = [worksheet.cell(row = curr_row, column = x).value for x in other_column_letter_inds]
        
        if cell_values == other_row_values:
            
            cell_value = worksheet.cell(row = curr_row, column = column_letter_ind).value
        
            break
            
        curr_row += 1
    if return_row: 
        return cell_value, curr_row - 1
    else: 
        return cell_value